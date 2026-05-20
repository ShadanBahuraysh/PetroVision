# ========================================================================================================
# PetroVision Analysis Routes
# --------------------------------------------------------------------------------------------------------
# This file contains all analysis-related API endpoints
# for the PetroVision system, including:
# - Running station performance analysis
# - Generating overview and ranking data
# - Explaining station results
# - Comparing stations
# - Exporting analysis reports (CSV / Excel)
# - Saving reports to the database
# - Clearing cached analysis data
#
# It also includes helper functions for building
# report summaries, insights, recommendations,
# and exporting formatted analysis reports.
# =========================================================================================================
import csv
import io
from datetime import datetime
from zoneinfo import ZoneInfo

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.supabase_client import supabase
from app.services.analysis_service import AnalysisService
from app.services.explanation_service import ExplanationService

router = APIRouter()

analysis_service = AnalysisService()
explanation_service = ExplanationService(analysis_service)


def _get_model_recommendation(station_id: str) -> str:
    fallback = "Review station performance and monitor operational indicators."
    try:
        station_analysis = analysis_service.get_station_analysis(station_id)
        actions = station_analysis.get("recommended_actions", [])
        if actions:
            first_action = actions[0]
            return first_action.get("action") or fallback
        return fallback
    except KeyError as action_error:
        print(f"Could not load recommendation for {station_id}: {action_error}")
        return fallback
    except ValueError as action_error:
        print(f"Could not load recommendation for {station_id}: {action_error}")
        return fallback
    except Exception as action_error:
        print(f"Could not load recommendation for {station_id}: {action_error}")
        return fallback


def _build_report_summary(rows):
    if not rows:
        return {
            "total_stations": 0,
            "average_score": 0,
            "good_stations": 0,
            "fair_stations": 0,
            "poor_stations": 0,
        }
    total = len(rows)
    average_score = round(sum(float(row.get("score") or 0) for row in rows) / total, 2)
    return {
        "total_stations": total,
        "average_score": average_score,
        "good_stations": len([r for r in rows if r.get("status") == "Good"]),
        "fair_stations": len([r for r in rows if r.get("status") == "Fair"]),
        "poor_stations": len([r for r in rows if r.get("status") == "Poor"]),
    }


def _build_report_insights(rows):
    if not rows:
        return []
    top_station = max(rows, key=lambda x: float(x.get("score") or 0))
    weakest_station = min(rows, key=lambda x: float(x.get("score") or 0))
    issue_counts = {}
    for row in rows:
        issue = row.get("top_issue") or "No major issue"
        issue_counts[issue] = issue_counts.get(issue, 0) + 1
    most_common_issue = max(issue_counts, key=issue_counts.get)
    return [
        f"Top performing station is {top_station['station_id']} with score {top_station['score']}.",
        f"Weakest station is {weakest_station['station_id']} with score {weakest_station['score']}.",
        f"Most common issue across stations is {most_common_issue}.",
    ]


def _build_export_rows():
    ranking = analysis_service.get_ranking()
    if not ranking:
        raise HTTPException(
            status_code=400,
            detail="No analysis data available. Please run analysis first.",
        )
    generated_at = datetime.now(ZoneInfo("Asia/Riyadh")).isoformat()
    rows = []
    for item in ranking:
        station_id = item.get("station_id")
        score = item.get("final_station_score")
        status = item.get("performance_status")
        priority = item.get("priority")
        top_issue = item.get("top_issue") or "No major issue"
        recommendation = item.get("recommendation") or "Review station performance and monitor operational indicators."
        rows.append({
            "station_id": station_id,
            "score": score,
            "status": status,
            "priority": priority,
            "top_issue": top_issue,
            "recommendation": recommendation,
        })
    summary = _build_report_summary(rows)
    insights = _build_report_insights(rows)
    return generated_at, summary, insights, rows


def _save_rows_to_report_table(rows, generated_at):
    if not rows:
        raise HTTPException(status_code=400, detail="No report rows available to save.")
    timestamp = datetime.now(ZoneInfo("Asia/Riyadh")).strftime("%Y%m%d%H%M%S")
    db_rows = []
    for index, row in enumerate(rows, start=1):
        db_rows.append({
            "report_id": f"RPT-{timestamp}-{index:04d}",
            "station_id": row.get("station_id"),
            "model_name": "PetroVision Analysis",
            "summary": (
                f"Station {row.get('station_id')} scored {row.get('score')} "
                f"and is classified as {row.get('status')}."
            ),
            "metric": "final_station_score",
            "detail": f"Priority: {row.get('priority')}. Top issue: {row.get('top_issue')}.",
            "recommendation": row.get("recommendation"),
            "generation_time": generated_at,
        })
    result = supabase.table("report").insert(db_rows).execute()
    if not result.data:
        raise HTTPException(status_code=500, detail="Report could not be saved to database.")
    return len(db_rows)


# ─────────────────────────────────────────────────────────────
# Run Analysis
# ─────────────────────────────────────────────────────────────

@router.get("/run-all")
def run_all_analysis(force: bool = False):
    try:
        return analysis_service.run_full_analysis(force=force)

    except MemoryError:
        raise HTTPException(
            status_code=500,
            detail="Not enough memory to load station data for analysis."
        )

    except FileNotFoundError as e:
        raise HTTPException(
            status_code=500,
            detail=f"ML model file not found: {str(e)}"
        )

    except RuntimeError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Analysis pipeline failed during execution: {str(e)}"
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while running the analysis."
        )


# ─────────────────────────────────────────────────────────────
# Overview
# ─────────────────────────────────────────────────────────────

@router.get("/overview")
def get_overview():
    try:
        return analysis_service.get_overview()

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail="Overview data not found. Please run analysis first."
        )

    except ZeroDivisionError:
        raise HTTPException(
            status_code=500,
            detail="Overview calculation failed: no station records to aggregate."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while loading the overview."
        )


# ─────────────────────────────────────────────────────────────
# Ranking
# ─────────────────────────────────────────────────────────────

@router.get("/ranking")
def get_ranking():
    try:
        return analysis_service.get_ranking()

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail="Ranking data not found. Please run analysis first."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while loading the station ranking."
        )


# ─────────────────────────────────────────────────────────────
# Top / Bottom 10
# ─────────────────────────────────────────────────────────────

@router.get("/top-bottom")
def get_top_bottom():
    try:
        return analysis_service.get_top_bottom_10()

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail="Top/bottom data not found. Please run analysis first."
        )

    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Not enough stations to produce a top/bottom ranking: {str(e)}"
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while loading top/bottom stations."
        )


# ─────────────────────────────────────────────────────────────
# Single Station Analysis
# ─────────────────────────────────────────────────────────────

@router.get("/station/{station_id}")
def analyze_station(station_id: str):
    try:
        return analysis_service.get_station_analysis(station_id)

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail=f"No analysis data found for station '{station_id}'."
        )

    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid station ID format: {str(e)}"
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail=f"An unexpected error occurred while analysing station '{station_id}'."
        )


# ─────────────────────────────────────────────────────────────
# AI Explanation - Single Station
# ─────────────────────────────────────────────────────────────

@router.get("/explain/station/{station_id}")
def explain_station(station_id: str):
    try:
        return explanation_service.explain_station(station_id)

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail=f"No analysis data available to explain for station '{station_id}'. Run analysis first."
        )

    except ConnectionError:
        raise HTTPException(
            status_code=503,
            detail="Could not reach the AI explanation service. A local explanation was returned instead."
        )

    except TimeoutError:
        raise HTTPException(
            status_code=504,
            detail="The AI explanation service timed out. Please try again."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail=f"An unexpected error occurred while generating the explanation for station '{station_id}'."
        )


# ─────────────────────────────────────────────────────────────
# AI Explanation - Network Overview
# ─────────────────────────────────────────────────────────────

@router.get("/explain/overview")
def explain_overview():
    try:
        return explanation_service.explain_overview()

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail="No overview data available to explain. Please run analysis first."
        )

    except ConnectionError:
        raise HTTPException(
            status_code=503,
            detail="Could not reach the AI explanation service. A local explanation was returned instead."
        )

    except TimeoutError:
        raise HTTPException(
            status_code=504,
            detail="The AI explanation service timed out. Please try again."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while generating the network overview explanation."
        )


# ─────────────────────────────────────────────────────────────
# Station Comparison
# ─────────────────────────────────────────────────────────────

@router.get("/compare/{station_1}/{station_2}")
def compare_stations(station_1: str, station_2: str):
    try:
        if station_1.strip().lower() == station_2.strip().lower():
            raise HTTPException(
                status_code=400,
                detail="Cannot compare a station with itself. Please provide two different station IDs."
            )

        return explanation_service.compare_stations(station_1, station_2)

    except HTTPException:
        raise

    except LookupError:
        raise HTTPException(
            status_code=404,
            detail=f"One or both stations ('{station_1}', '{station_2}') have no analysis data. Run analysis first."
        )

    except ConnectionError:
        raise HTTPException(
            status_code=503,
            detail="Could not reach the AI explanation service. A local comparison was returned instead."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail=f"An unexpected error occurred while comparing stations '{station_1}' and '{station_2}'."
        )


# ─────────────────────────────────────────────────────────────
# Export - JSON Payload
# ─────────────────────────────────────────────────────────────

@router.get("/export")
def export_analysis_report(save_to_db: bool = False):
    try:
        generated_at, summary, insights, rows = _build_export_rows()
        saved_to_db = False

        if save_to_db:
            _save_rows_to_report_table(rows, generated_at)
            saved_to_db = True

        return {
            "generated_at": generated_at,
            "saved_to_db": saved_to_db,
            "summary": summary,
            "insights": insights,
            "rows": rows,
        }

    except HTTPException:
        raise

    except PermissionError:
        raise HTTPException(
            status_code=403,
            detail="Permission denied while accessing report data."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while building the export report."
        )


# ─────────────────────────────────────────────────────────────
# Export - File Download (CSV / Excel)
# ─────────────────────────────────────────────────────────────

@router.get("/export/download")
def download_analysis_report(file_type: str = "csv"):
    try:
        generated_at, summary, insights, rows = _build_export_rows()

        if not rows:
            raise HTTPException(status_code=400, detail="No report data available.")

        normalized_file_type = file_type.strip().lower()

        if normalized_file_type in ["excel", "xlsx"]:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Analysis Report"

            sheet["A1"] = "PetroVision Analysis Report"
            sheet["A1"].font = Font(bold=True, size=16, color="1A2E35")
            sheet["A2"] = f"Generated at: {generated_at}"
            sheet["A2"].font = Font(color="4195AF")

            sheet.append([])
            sheet.append(["Summary"])
            sheet.append(["Total Stations", summary["total_stations"]])
            sheet.append(["Average Score", summary["average_score"]])
            sheet.append(["Good Stations", summary["good_stations"]])
            sheet.append(["Fair Stations", summary["fair_stations"]])
            sheet.append(["Poor Stations", summary["poor_stations"]])

            sheet.append([])
            sheet.append(["Key Insights"])
            for insight in insights:
                sheet.append([insight])

            sheet.append([])
            headers = list(rows[0].keys())
            sheet.append(headers)

            header_row = sheet.max_row
            for cell in sheet[header_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1A2E35")
                cell.alignment = Alignment(horizontal="center")

            for row in rows:
                sheet.append([row.get(header, "") for header in headers])

            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))
                sheet.column_dimensions[column_letter].width = min(max_length + 3, 45)

            stream = io.BytesIO()
            workbook.save(stream)
            stream.seek(0)

            return StreamingResponse(
                stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=petrovision_analysis_report.xlsx"
                },
            )

        if normalized_file_type != "csv":
            raise HTTPException(status_code=400, detail="Unsupported file type. Use 'csv' or 'excel'.")

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["PetroVision Analysis Report"])
        writer.writerow(["Generated at", generated_at])
        writer.writerow([])
        writer.writerow(["Summary"])
        for key, value in summary.items():
            writer.writerow([key, value])

        writer.writerow([])
        writer.writerow(["Key Insights"])
        for insight in insights:
            writer.writerow([insight])

        writer.writerow([])
        headers = list(rows[0].keys())
        writer.writerow(headers)

        for row in rows:
            writer.writerow([row.get(header, "") for header in headers])

        csv_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        csv_bytes.seek(0)

        return StreamingResponse(
            csv_bytes,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=petrovision_analysis_report.csv"
            },
        )

    except HTTPException:
        raise

    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type or report data: {str(e)}"
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while generating the download file."
        )


# ─────────────────────────────────────────────────────────────
# Save Report to Database
# ─────────────────────────────────────────────────────────────

@router.post("/export/save")
def save_analysis_report():
    try:
        generated_at, summary, insights, rows = _build_export_rows()
        saved_count = _save_rows_to_report_table(rows, generated_at)

        return {
            "message": "Report saved successfully",
            "saved_to_db": True,
            "rows_count": saved_count,
            "summary": summary,
            "insights": insights,
        }

    except HTTPException:
        raise

    except ConnectionError:
        raise HTTPException(
            status_code=503,
            detail="Could not connect to the database to save the report. Please try again."
        )

    except PermissionError:
        raise HTTPException(
            status_code=403,
            detail="Permission denied when writing the report to the database."
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while saving the report to the database."
        )


# ─────────────────────────────────────────────────────────────
# Clear Cache
# ─────────────────────────────────────────────────────────────

@router.get("/clear-cache")
def clear_cache():
    try:
        analysis_result = analysis_service.clear_cache()
        explanation_service.clear_explanation_cache()

        return {
            "message": "All caches cleared successfully",
            "analysis": analysis_result,
        }

    except AttributeError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Cache clearing failed - service not properly initialised: {str(e)}"
        )

    except Exception:
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while clearing the analysis cache."
        )